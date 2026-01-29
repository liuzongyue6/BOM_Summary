"""
BOM Comparison Module
Author: Zongyue Liu
Date: 2026-01-28

Standalone module for comparing two BOM Excel files and generating detailed comparison reports.
Can be used both as a CLI tool and programmatically (e.g., by Streamlit).

Features:
- Categorizes parts into 4 types: NoChange, Part_Rev_NoChange, RevChange, PartChange
- Side-by-side format: File1 columns | File2 columns
- Yellow highlighting for differences
- Summary sheet combining all categories

Usage:
    1. CLI mode:
        python bom_comparison.py --file1 path/to/file1.xlsx --file2 path/to/file2.xlsx
    
    2. Programmatic mode (with file bytes):
        comparator = BOMComparator(file1_bytes, file1_name, file2_bytes, file2_name)
        success, message = comparator.load_data()
        success, message = comparator.compare()
        buffer = comparator.generate_comparison_report()
    
    3. Programmatic mode (with file paths):
        comparator = BOMComparator.from_files(file1_path, file2_path)
        success, message = comparator.process()
        buffer = comparator.generate_comparison_report()
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional, Callable, Set
from io import BytesIO
from enum import Enum
from dataclasses import dataclass
import datetime
import argparse
import os


class ChangeCategory(Enum):
    """Enumeration of change categories"""
    NO_CHANGE = "NoChange"
    PART_REV_NO_CHANGE = "Part_Rev_NoChange"
    REV_CHANGE = "RevChange"
    PART_CHANGE = "PartChange"


@dataclass
class ComparisonPair:
    """Data structure for a comparison pair"""
    part_number1: str
    rev1: str
    data1: Dict[str, Any]
    part_number2: str
    rev2: str
    data2: Dict[str, Any]
    category: ChangeCategory
    
    @property
    def has_file1(self) -> bool:
        """Check if File1 data exists"""
        return bool(self.part_number1)
    
    @property
    def has_file2(self) -> bool:
        """Check if File2 data exists"""
        return bool(self.part_number2)


class BOMComparator:
    """Compare two BOM Excel files and generate detailed comparison report"""
    
    # Default columns to compare (can be overridden in __init__)
    DEFAULT_COMPARISON_COLUMNS = [
        'CAD OEM Part Number',
        'CAD OEM Rev',
        'Material Spec',
        'CAD Oem Name',
        'Thickness'
    ]
    
    def __init__(self, 
                 file1_bytes: bytes = None, 
                 file1_name: str = None, 
                 file2_bytes: bytes = None, 
                 file2_name: str = None,
                 comparison_columns: List[str] = None,
                 log_callback: Optional[Callable[[str], None]] = None):
        """
        Initialize BOM Comparator
        
        Args:
            file1_bytes: Bytes content of first Excel file (optional if using from_files)
            file1_name: Filename of first Excel file
            file2_bytes: Bytes content of second Excel file (optional if using from_files)
            file2_name: Filename of second Excel file
            comparison_columns: List of columns to compare (uses DEFAULT_COMPARISON_COLUMNS if None)
            log_callback: Optional callback function for real-time log updates (for Streamlit)
        """
        self.file1_name = Path(file1_name).stem if file1_name else "File1"
        self.file2_name = Path(file2_name).stem if file2_name else "File2"
        self.file1_full_name = file1_name or "File1"
        self.file2_full_name = file2_name or "File2"
        
        # Comparison columns
        self.comparison_columns = comparison_columns or self.DEFAULT_COMPARISON_COLUMNS
        
        # Log callback
        self.log_callback = log_callback
        self.log_messages = []
        self.start_time = datetime.datetime.now()
        
        # Load workbooks if bytes provided
        self.wb1 = None
        self.wb2 = None
        self.ws1 = None
        self.ws2 = None
        
        if file1_bytes and file2_bytes:
            self.wb1 = openpyxl.load_workbook(BytesIO(file1_bytes), data_only=True)
            self.wb2 = openpyxl.load_workbook(BytesIO(file2_bytes), data_only=True)
            self.ws1 = self.wb1.worksheets[0]
            self.ws2 = self.wb2.worksheets[0]
        
        # Data storage: key = (part_number, rev), value = dict of all columns
        self.data1: Dict[Tuple[str, str], Dict[str, Any]] = {}
        self.data2: Dict[Tuple[str, str], Dict[str, Any]] = {}
        
        # Index by part number only (for RevChange detection)
        self.data1_by_part: Dict[str, List[Tuple[str, str]]] = {}
        self.data2_by_part: Dict[str, List[Tuple[str, str]]] = {}
        
        # Comparison results: categorized pairs
        self.comparison_pairs: Dict[ChangeCategory, List[ComparisonPair]] = {
            ChangeCategory.NO_CHANGE: [],
            ChangeCategory.PART_REV_NO_CHANGE: [],
            ChangeCategory.REV_CHANGE: [],
            ChangeCategory.PART_CHANGE: []
        }
        
        # Styling
        self.yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        self.header_font = Font(bold=True)
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    @classmethod
    def from_files(cls, file1_path: str, file2_path: str, 
                   comparison_columns: List[str] = None,
                   log_callback: Optional[Callable[[str], None]] = None):
        """
        Create BOMComparator from file paths (for CLI and file-based usage)
        
        Args:
            file1_path: Path to first Excel file
            file2_path: Path to second Excel file
            comparison_columns: List of columns to compare
            log_callback: Optional callback function for log updates
        
        Returns:
            BOMComparator instance
        """
        # Read files
        with open(file1_path, 'rb') as f:
            file1_bytes = f.read()
        with open(file2_path, 'rb') as f:
            file2_bytes = f.read()
        
        file1_name = Path(file1_path).name
        file2_name = Path(file2_path).name
        
        return cls(file1_bytes, file1_name, file2_bytes, file2_name, 
                  comparison_columns, log_callback)
    
    def log(self, message: str, level: str = "INFO"):
        """Add message to log"""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}"
        self.log_messages.append(log_entry)
        
        # Call callback if provided (for Streamlit real-time updates)
        if self.log_callback:
            self.log_callback(log_entry)
        else:
            # Default: print to console (for CLI usage)
            print(log_entry)
    
    def sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """
        Sanitize sheet name to comply with Excel requirements
        
        Args:
            name: Sheet name to sanitize
            max_length: Maximum length (Excel limit is 31)
        
        Returns:
            Sanitized sheet name
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Truncate to max length
        if len(name) > max_length:
            name = name[:max_length]
        
        return name
    
    def extract_headers(self, worksheet: Any) -> Dict[str, int]:
        """
        Extract column headers from worksheet
        
        Args:
            worksheet: openpyxl worksheet
        
        Returns:
            Dictionary mapping column names to column indices
        """
        headers = {}
        for col_idx, cell in enumerate(worksheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        return headers
    
    def validate_columns(self, headers: Dict[str, int], filename: str) -> Tuple[bool, List[str]]:
        """
        Validate that required comparison columns exist
        
        Args:
            headers: Dictionary of column headers
            filename: Name of file being validated
        
        Returns:
            Tuple of (is_valid, missing_columns)
        """
        missing = []
        for col in self.comparison_columns:
            if col not in headers:
                missing.append(col)
        
        return len(missing) == 0, missing
    
    def load_data(self) -> Tuple[bool, str]:
        """
        Load data from both Excel files
        
        Returns:
            Tuple of (success, error_message)
        """
        try:
            self.log("Loading data from both Excel files...")
            
            # Extract headers
            headers1 = self.extract_headers(self.ws1)
            headers2 = self.extract_headers(self.ws2)
            
            self.log(f"File 1: Found {len(headers1)} column headers")
            self.log(f"File 2: Found {len(headers2)} column headers")
            
            # Validate columns
            valid1, missing1 = self.validate_columns(headers1, self.file1_name)
            valid2, missing2 = self.validate_columns(headers2, self.file2_name)
            
            if not valid1:
                error_msg = f"File 1 missing columns: {', '.join(missing1)}"
                self.log(error_msg, "ERROR")
                return False, error_msg
            if not valid2:
                error_msg = f"File 2 missing columns: {', '.join(missing2)}"
                self.log(error_msg, "ERROR")
                return False, error_msg
            
            # Load data from file 1
            self.log(f"Loading data from {self.file1_name}...")
            self.data1, self.data1_by_part = self._load_worksheet_data(self.ws1, headers1)
            self.log(f"✓ Loaded {len(self.data1)} parts from File 1")
            
            # Load data from file 2
            self.log(f"Loading data from {self.file2_name}...")
            self.data2, self.data2_by_part = self._load_worksheet_data(self.ws2, headers2)
            self.log(f"✓ Loaded {len(self.data2)} parts from File 2")
            
            return True, ""
        
        except Exception as e:
            error_msg = f"Error loading data: {str(e)}"
            self.log(error_msg, "ERROR")
            return False, error_msg
    
    def _load_worksheet_data(self, worksheet: Any, headers: Dict[str, int]) -> Tuple[Dict[Tuple[str, str], Dict[str, Any]], Dict[str, List[Tuple[str, str]]]]:
        """
        Load data from worksheet into dictionary
        
        Args:
            worksheet: openpyxl worksheet
            headers: Column header mapping
        
        Returns:
            Tuple of (data_dict, part_index)
            - data_dict: Dictionary keyed by (part_number, rev)
            - part_index: Dictionary keyed by part_number, values are list of (part_number, rev) tuples
        """
        data = {}
        part_index = {}
        
        for row_idx in range(2, worksheet.max_row + 1):
            # Get key columns
            part_num_col = headers.get('CAD OEM Part Number')
            rev_col = headers.get('CAD OEM Rev')
            
            if not part_num_col or not rev_col:
                continue
            
            part_num = worksheet.cell(row_idx, part_num_col).value
            rev = worksheet.cell(row_idx, rev_col).value
            
            # Skip rows without part number
            if not part_num:
                continue
            
            # Normalize keys
            part_num_str = str(part_num).strip() if part_num else ''
            rev_str = str(rev).strip() if rev else ''
            key = (part_num_str, rev_str)
            
            # Extract all comparison columns
            row_data = {}
            for col_name in self.comparison_columns:
                col_idx = headers.get(col_name)
                if col_idx:
                    cell_value = worksheet.cell(row_idx, col_idx).value
                    # Convert to string and strip, preserve None as empty string
                    row_data[col_name] = str(cell_value).strip() if cell_value is not None else ''
            
            # Store data (if duplicate keys exist, last one wins)
            data[key] = row_data
            
            # Build part number index
            if part_num_str not in part_index:
                part_index[part_num_str] = []
            if key not in part_index[part_num_str]:
                part_index[part_num_str].append(key)
        
        return data, part_index
    
    def compare(self) -> Tuple[bool, str]:
        """
        Compare the two BOM files and categorize parts
        
        Returns:
            Tuple of (success, message)
        """
        try:
            self.log("Comparing BOM files and categorizing parts...")
            
            # Clear previous results
            for category in ChangeCategory:
                self.comparison_pairs[category] = []
            
            # Track processed keys to avoid duplicates
            processed_keys1: Set[Tuple[str, str]] = set()
            processed_keys2: Set[Tuple[str, str]] = set()
            
            # Step 1: Find exact matches (Part Number + Rev)
            keys1 = set(self.data1.keys())
            keys2 = set(self.data2.keys())
            matching_keys = keys1.intersection(keys2)
            
            for key in matching_keys:
                part_num, rev = key
                data1 = self.data1[key]
                data2 = self.data2[key]
                
                # Check if all other columns are identical
                other_columns_identical = self._check_all_columns_identical(data1, data2)
                
                if other_columns_identical:
                    # NoChange: Everything is identical
                    pair = ComparisonPair(
                        part_number1=part_num,
                        rev1=rev,
                        data1=data1,
                        part_number2=part_num,
                        rev2=rev,
                        data2=data2,
                        category=ChangeCategory.NO_CHANGE
                    )
                    self.comparison_pairs[ChangeCategory.NO_CHANGE].append(pair)
                else:
                    # Part_Rev_NoChange: Part + Rev same, but other columns differ
                    pair = ComparisonPair(
                        part_number1=part_num,
                        rev1=rev,
                        data1=data1,
                        part_number2=part_num,
                        rev2=rev,
                        data2=data2,
                        category=ChangeCategory.PART_REV_NO_CHANGE
                    )
                    self.comparison_pairs[ChangeCategory.PART_REV_NO_CHANGE].append(pair)
                
                processed_keys1.add(key)
                processed_keys2.add(key)
            
            # Step 2: Find RevChange (same Part Number, different Rev)
            # Check remaining parts in File1
            remaining_keys1 = keys1 - processed_keys1
            for key1 in remaining_keys1:
                part_num1, rev1 = key1
                
                # Check if this part number exists in File2 with different rev
                if part_num1 in self.data2_by_part:
                    # Find best match (first unprocessed rev)
                    for key2 in self.data2_by_part[part_num1]:
                        if key2 not in processed_keys2:
                            part_num2, rev2 = key2
                            data1 = self.data1[key1]
                            data2 = self.data2[key2]
                            
                            pair = ComparisonPair(
                                part_number1=part_num1,
                                rev1=rev1,
                                data1=data1,
                                part_number2=part_num2,
                                rev2=rev2,
                                data2=data2,
                                category=ChangeCategory.REV_CHANGE
                            )
                            self.comparison_pairs[ChangeCategory.REV_CHANGE].append(pair)
                            
                            processed_keys1.add(key1)
                            processed_keys2.add(key2)
                            break
            
            # Step 3: PartChange - Parts that don't match
            # Remaining File1 parts (not found in File2)
            remaining_keys1 = keys1 - processed_keys1
            for key1 in remaining_keys1:
                part_num1, rev1 = key1
                data1 = self.data1[key1]
                
                pair = ComparisonPair(
                    part_number1=part_num1,
                    rev1=rev1,
                    data1=data1,
                    part_number2='',
                    rev2='',
                    data2={col: '' for col in self.comparison_columns},
                    category=ChangeCategory.PART_CHANGE
                )
                self.comparison_pairs[ChangeCategory.PART_CHANGE].append(pair)
            
            # Remaining File2 parts (not found in File1)
            remaining_keys2 = keys2 - processed_keys2
            for key2 in remaining_keys2:
                part_num2, rev2 = key2
                data2 = self.data2[key2]
                
                pair = ComparisonPair(
                    part_number1='',
                    rev1='',
                    data1={col: '' for col in self.comparison_columns},
                    part_number2=part_num2,
                    rev2=rev2,
                    data2=data2,
                    category=ChangeCategory.PART_CHANGE
                )
                self.comparison_pairs[ChangeCategory.PART_CHANGE].append(pair)
            
            # Log results
            self.log(f"✓ NoChange: {len(self.comparison_pairs[ChangeCategory.NO_CHANGE])} parts")
            self.log(f"✓ Part_Rev_NoChange: {len(self.comparison_pairs[ChangeCategory.PART_REV_NO_CHANGE])} parts")
            self.log(f"✓ RevChange: {len(self.comparison_pairs[ChangeCategory.REV_CHANGE])} parts")
            self.log(f"✓ PartChange: {len(self.comparison_pairs[ChangeCategory.PART_CHANGE])} parts")
            
            total_pairs = sum(len(pairs) for pairs in self.comparison_pairs.values())
            message = f"Comparison complete: {total_pairs} total comparisons across 4 categories"
            
            return True, message
        
        except Exception as e:
            error_msg = f"Error during comparison: {str(e)}"
            self.log(error_msg, "ERROR")
            return False, error_msg
    
    def _check_all_columns_identical(self, data1: Dict[str, Any], data2: Dict[str, Any]) -> bool:
        """
        Check if all columns (except Part Number and Rev) are identical
        
        Args:
            data1: Data from File1
            data2: Data from File2
        
        Returns:
            True if all columns are identical
        """
        exclude_cols = {'CAD OEM Part Number', 'CAD OEM Rev'}
        
        for col in self.comparison_columns:
            if col in exclude_cols:
                continue
            
            val1 = str(data1.get(col, '')).strip()
            val2 = str(data2.get(col, '')).strip()
            
            if val1 != val2:
                return False
        
        return True
    
    def _values_differ(self, val1: Any, val2: Any) -> bool:
        """
        Check if two values are different
        
        Args:
            val1: First value
            val2: Second value
        
        Returns:
            True if values differ
        """
        str1 = str(val1).strip() if val1 is not None else ''
        str2 = str(val2).strip() if val2 is not None else ''
        return str1 != str2
    
    def generate_comparison_report(self) -> BytesIO:
        """
        Generate Excel comparison report with 5 sheets
        
        Returns:
            BytesIO buffer containing the Excel file
        """
        self.log("Generating comparison report...")
        
        # Create new workbook
        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)
        
        # Sheet 1: NoChange
        self._create_category_sheet(output_wb, ChangeCategory.NO_CHANGE)
        self.log("✓ Created 'NoChange' sheet")
        
        # Sheet 2: Part_Rev_NoChange
        self._create_category_sheet(output_wb, ChangeCategory.PART_REV_NO_CHANGE)
        self.log("✓ Created 'Part_Rev_NoChange' sheet")
        
        # Sheet 3: RevChange
        self._create_category_sheet(output_wb, ChangeCategory.REV_CHANGE)
        self.log("✓ Created 'RevChange' sheet")
        
        # Sheet 4: PartChange
        self._create_category_sheet(output_wb, ChangeCategory.PART_CHANGE)
        self.log("✓ Created 'PartChange' sheet")
        
        # Sheet 5: Summary
        self._create_summary_sheet(output_wb)
        self.log("✓ Created 'Summary' sheet")
        
        # Save to BytesIO
        buffer = BytesIO()
        output_wb.save(buffer)
        buffer.seek(0)
        
        self.log("✓ Comparison report generated successfully")
        
        return buffer
    
    def _create_category_sheet(self, workbook: Any, category: ChangeCategory):
        """
        Create a sheet for a specific category with side-by-side format
        
        Args:
            workbook: openpyxl workbook
            category: Category to create sheet for
        """
        ws = workbook.create_sheet(title=category.value)
        pairs = self.comparison_pairs[category]
        
        # Create headers: File1 columns | File2 columns
        col_idx = 1
        for col_name in self.comparison_columns:
            # File1 column header
            cell = ws.cell(1, col_idx, f"File1_{col_name}")
            cell.fill = self.header_fill
            cell.font = self.header_font
            col_idx += 1
        
        for col_name in self.comparison_columns:
            # File2 column header
            cell = ws.cell(1, col_idx, f"File2_{col_name}")
            cell.fill = self.header_fill
            cell.font = self.header_font
            col_idx += 1
        
        # Write data
        row_idx = 2
        for pair in pairs:
            col_idx = 1
            
            # File1 data
            for col_name in self.comparison_columns:
                value1 = pair.data1.get(col_name, '')
                cell = ws.cell(row_idx, col_idx, value1)
                
                # Highlight if differs from File2
                if pair.has_file2:
                    value2 = pair.data2.get(col_name, '')
                    if self._values_differ(value1, value2):
                        cell.fill = self.yellow_fill
                
                col_idx += 1
            
            # File2 data
            for col_name in self.comparison_columns:
                value2 = pair.data2.get(col_name, '')
                cell = ws.cell(row_idx, col_idx, value2)
                
                # Highlight if differs from File1
                if pair.has_file1:
                    value1 = pair.data1.get(col_name, '')
                    if self._values_differ(value1, value2):
                        cell.fill = self.yellow_fill
                
                col_idx += 1
            
            row_idx += 1
    
    def _create_summary_sheet(self, workbook: Any):
        """
        Create summary sheet combining all categories
        
        Args:
            workbook: openpyxl workbook
        """
        ws = workbook.create_sheet(title="Summary")
        
        # Create headers
        col_idx = 1
        for col_name in self.comparison_columns:
            cell = ws.cell(1, col_idx, f"File1_{col_name}")
            cell.fill = self.header_fill
            cell.font = self.header_font
            col_idx += 1
        
        for col_name in self.comparison_columns:
            cell = ws.cell(1, col_idx, f"File2_{col_name}")
            cell.fill = self.header_fill
            cell.font = self.header_font
            col_idx += 1
        
        row_idx = 2
        
        # Add data from each category with section headers and spacing
        for category in [ChangeCategory.NO_CHANGE, ChangeCategory.PART_REV_NO_CHANGE, 
                        ChangeCategory.REV_CHANGE, ChangeCategory.PART_CHANGE]:
            pairs = self.comparison_pairs[category]
            
            if pairs:
                # Section header
                cell = ws.cell(row_idx, 1, f"--- {category.value} ({len(pairs)} items) ---")
                cell.font = Font(bold=True, size=12)
                row_idx += 1
                
                # Add pairs
                for pair in pairs:
                    col_idx = 1
                    
                    # File1 data
                    for col_name in self.comparison_columns:
                        value1 = pair.data1.get(col_name, '')
                        cell = ws.cell(row_idx, col_idx, value1)
                        
                        if pair.has_file2:
                            value2 = pair.data2.get(col_name, '')
                            if self._values_differ(value1, value2):
                                cell.fill = self.yellow_fill
                        
                        col_idx += 1
                    
                    # File2 data
                    for col_name in self.comparison_columns:
                        value2 = pair.data2.get(col_name, '')
                        cell = ws.cell(row_idx, col_idx, value2)
                        
                        if pair.has_file1:
                            value1 = pair.data1.get(col_name, '')
                            if self._values_differ(value1, value2):
                                cell.fill = self.yellow_fill
                        
                        col_idx += 1
                    
                    row_idx += 1
                
                # Blank row between sections
                row_idx += 1
    
    def get_stats(self) -> Dict[str, int]:
        """
        Get comparison statistics
        
        Returns:
            Dictionary with comparison statistics
        """
        return {
            'total_file1': len(self.data1),
            'total_file2': len(self.data2),
            'no_change': len(self.comparison_pairs[ChangeCategory.NO_CHANGE]),
            'part_rev_no_change': len(self.comparison_pairs[ChangeCategory.PART_REV_NO_CHANGE]),
            'rev_change': len(self.comparison_pairs[ChangeCategory.REV_CHANGE]),
            'part_change': len(self.comparison_pairs[ChangeCategory.PART_CHANGE])
        }
    
    def write_log_file(self, log_file_path: str):
        """
        Write all log messages to file
        
        Args:
            log_file_path: Path where to save the log file
        """
        with open(log_file_path, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("BOM COMPARISON LOG\n")
            f.write("="*80 + "\n\n")
            f.write(f"Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"File 1: {self.file1_full_name}\n")
            f.write(f"File 2: {self.file2_full_name}\n")
            f.write("\n" + "="*80 + "\n")
            f.write("COMPARISON LOG\n")
            f.write("="*80 + "\n\n")
            
            for msg in self.log_messages:
                f.write(msg + "\n")
            
            f.write("\n" + "="*80 + "\n")
            f.write("COMPARISON STATISTICS\n")
            f.write("="*80 + "\n\n")
            
            stats = self.get_stats()
            f.write(f"File 1 total parts: {stats['total_file1']}\n")
            f.write(f"File 2 total parts: {stats['total_file2']}\n")
            f.write(f"\nCategorization:\n")
            f.write(f"  NoChange: {stats['no_change']}\n")
            f.write(f"  Part_Rev_NoChange: {stats['part_rev_no_change']}\n")
            f.write(f"  RevChange: {stats['rev_change']}\n")
            f.write(f"  PartChange: {stats['part_change']}\n")
            
            end_time = datetime.datetime.now()
            duration = end_time - self.start_time
            f.write(f"\nEnd Time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Duration: {duration.total_seconds():.2f} seconds\n")
            f.write("\n" + "="*80 + "\n")
            f.write("COMPARISON COMPLETED SUCCESSFULLY\n")
            f.write("="*80 + "\n")
    
    def process(self, output_file: str = None, log_file: str = None) -> bool:
        """
        Main processing pipeline (for CLI and file-based usage)
        
        Args:
            output_file: Path where to save comparison report (auto-generated if None)
            log_file: Path where to save log file (auto-generated if None)
        
        Returns:
            True if successful, False otherwise
        """
        try:
            self.log("="*80)
            self.log("BOM COMPARISON STARTED")
            self.log("="*80)
            
            # Step 1: Load data
            success, message = self.load_data()
            if not success:
                return False
            
            # Step 2: Compare
            success, message = self.compare()
            if not success:
                return False
            
            # Step 3: Generate report
            buffer = self.generate_comparison_report()
            
            # Step 4: Save to file
            if output_file is None:
                current_date = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"comparison_result_{current_date}.xlsx"
            
            self.log(f"Saving comparison report to: {output_file}")
            with open(output_file, 'wb') as f:
                f.write(buffer.getvalue())
            self.log("✓ Comparison report saved successfully")
            
            # Step 5: Write log file
            if log_file is None:
                current_date = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                log_file = f"comparison_log_{current_date}.txt"
            
            self.log(f"Writing log file to: {log_file}")
            self.write_log_file(log_file)
            
            self.log("="*80)
            self.log("BOM COMPARISON COMPLETED SUCCESSFULLY")
            self.log("="*80)
            
            return True
            
        except Exception as e:
            self.log(f"FATAL ERROR: {str(e)}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            
            # Try to write log even on failure
            if log_file:
                try:
                    self.write_log_file(log_file)
                except:
                    pass
            
            return False


def main():
    """Main entry point for CLI usage"""
    parser = argparse.ArgumentParser(
        description='Compare two BOM Excel files and generate detailed comparison report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python bom_comparison.py --file1 data/BOM1.xlsx --file2 data/BOM2.xlsx
  python bom_comparison.py --file1 BOM1.xlsx --file2 BOM2.xlsx --output result.xlsx
  python bom_comparison.py --file1 BOM1.xlsx --file2 BOM2.xlsx --output result.xlsx --log comparison.log

Output Format:
  The comparison report contains 5 sheets:
  1. NoChange: Parts that are completely identical
  2. Part_Rev_NoChange: Same Part+Rev, but other columns differ
  3. RevChange: Same Part Number, different Rev
  4. PartChange: Parts found in only one file
  5. Summary: All categories combined with section headers
  
  All sheets use side-by-side format: File1 columns | File2 columns
  Differences are highlighted in yellow.
        """
    )
    
    parser.add_argument('--file1', required=True, help='Path to first BOM Excel file')
    parser.add_argument('--file2', required=True, help='Path to second BOM Excel file')
    parser.add_argument('--output', '-o', help='Output comparison report path (default: auto-generated)')
    parser.add_argument('--log', '-l', help='Log file path (default: auto-generated)')
    
    args = parser.parse_args()
    
    # Verify input files exist
    if not os.path.exists(args.file1):
        print(f"ERROR: File 1 not found: {args.file1}")
        return 1
    
    if not os.path.exists(args.file2):
        print(f"ERROR: File 2 not found: {args.file2}")
        return 1
    
    print(f"File 1: {args.file1}")
    print(f"File 2: {args.file2}")
    print()
    
    # Create comparator and process
    comparator = BOMComparator.from_files(args.file1, args.file2)
    success = comparator.process(output_file=args.output, log_file=args.log)
    
    if success:
        print()
        print("✓ Comparison complete!")
        stats = comparator.get_stats()
        print(f"  NoChange: {stats['no_change']}")
        print(f"  Part_Rev_NoChange: {stats['part_rev_no_change']}")
        print(f"  RevChange: {stats['rev_change']}")
        print(f"  PartChange: {stats['part_change']}")
        return 0
    else:
        print()
        print("✗ Comparison failed!")
        return 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
