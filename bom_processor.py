"""
BOM Processor - Extract and compress hierarchical BOM data from Excel files
Author: Zongyue Liu
Date: 2026-01-24
"""

import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
import datetime
import os
from typing import Dict, List, Tuple, Any

# 导入配置
from bom_config import BOMConfig


class BOMProcessor:
    """Process hierarchical BOM Excel files"""
    
    # 类级别配置：从BOMConfig导入
    CONFIG = BOMConfig
    
    def __init__(self, input_file: str = None, output_file: str = None, 
                 summary_file: str = None, log_file: str = None, 
                 log_callback=None, config=None):
        """
        初始化BOM处理器
        
        Args:
            config: 可选的配置对象，用于覆盖默认配置
        """
        self.input_file = input_file
        self.output_file = output_file
        self.summary_file = summary_file
        self.log_file = log_file
        self.log_callback = log_callback
        self.log_messages = []
        self.start_time = datetime.datetime.now()
        
        # 允许自定义配置
        if config:
            self.CONFIG = config
        
        # Statistics
        self.stats = {
            'level1_count': 0,
            'level2_count': 0,
            'sheets_created': [],
            'rows_per_sheet': {},
            'dedup_stats': {}
        }
    
    def log(self, message: str, level: str = "INFO"):
        """Add message to log"""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = f"[{timestamp}] [{level}] {message}"
        self.log_messages.append(log_entry)
        
        if self.log_callback:
            self.log_callback(log_entry)
        else:
            print(log_entry)
    
    def write_log_file(self):
        """Write all log messages to file"""
        with open(self.log_file, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("BOM PROCESSING LOG\n")
            f.write("="*80 + "\n\n")
            f.write(f"Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Input File: {self.input_file}\n")
            f.write(f"Output File: {self.output_file}\n")
            f.write("\n" + "="*80 + "\n")
            f.write("PROCESSING LOG\n")
            f.write("="*80 + "\n\n")
            
            for msg in self.log_messages:
                f.write(msg + "\n")
            
            f.write("\n" + "="*80 + "\n")
            f.write("STATISTICS SUMMARY\n")
            f.write("="*80 + "\n\n")
            f.write(f"Level 1 items found: {self.stats['level1_count']}\n")
            f.write(f"Level 2 items found: {self.stats['level2_count']}\n")
            f.write(f"Sheets created: {len(self.stats['sheets_created'])}\n")
            f.write(f"Sheet names: {', '.join(self.stats['sheets_created'])}\n\n")
            
            f.write("Rows per sheet:\n")
            for sheet_name, count in self.stats['rows_per_sheet'].items():
                f.write(f"  {sheet_name}: {count} rows\n")
            
            if self.stats['dedup_stats']:
                f.write("\nDeduplication statistics:\n")
                for sheet_name, stats in self.stats['dedup_stats'].items():
                    f.write(f"  {sheet_name}:\n")
                    f.write(f"    Original rows: {stats['original']}\n")
                    f.write(f"    After dedup: {stats['deduplicated']}\n")
                    f.write(f"    Removed: {stats['removed']}\n")
            
            end_time = datetime.datetime.now()
            duration = end_time - self.start_time
            f.write(f"\nEnd Time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Duration: {duration.total_seconds():.2f} seconds\n")
            f.write("\n" + "="*80 + "\n")
            f.write("PROCESSING COMPLETED SUCCESSFULLY\n")
            f.write("="*80 + "\n")
            
    def load_workbook_data(self) -> Tuple[Any, Any, Dict[str, int], List[Dict]]:
        """Load workbook and extract hierarchical data with outline levels"""
        self.log(f"Loading workbook: {self.input_file}")
        
        try:
            wb = load_workbook(self.input_file, data_only=True, keep_vba=False)
            self.log(f"Workbook loaded successfully. Sheets: {wb.sheetnames}")
        except Exception as e:
            self.log(f"Error loading workbook: {e}", "ERROR")
            raise
        
        main_sheet = wb.worksheets[0]
        sheet_name = main_sheet.title
        self.log(f"Processing main sheet: '{sheet_name}' ({main_sheet.max_row} rows, {main_sheet.max_column} cols)")
        
        # Extract headers
        headers = {}
        for col_idx in range(1, main_sheet.max_column + 1):
            cell_value = main_sheet.cell(1, col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx
        
        self.log(f"Found {len(headers)} column headers")
        
        # 🔧 使用配置验证必需列
        is_valid, missing_cols = self.CONFIG.validate_columns(headers.keys())
        if not is_valid:
            self.log(f"WARNING: Missing required columns: {missing_cols}", "WARNING")
        
        # Extract data rows
        data_rows = []
        for row_idx in range(2, main_sheet.max_row + 1):
            row_dim = main_sheet.row_dimensions[row_idx]
            outline_level = row_dim.outline_level if row_dim.outline_level else 0
            
            row_data = {'_row_num': row_idx, '_outline_level': outline_level}
            
            for col_name, col_idx in headers.items():
                cell_value = main_sheet.cell(row_idx, col_idx).value
                row_data[col_name] = cell_value
            
            data_rows.append(row_data)
        
        self.log(f"Extracted {len(data_rows)} data rows")
        
        return wb, main_sheet, headers, data_rows
    
    def analyze_hierarchy(self, data_rows: List[Dict]) -> Tuple[List[Dict], List[Dict]]:
        """Analyze and extract Level 1 and Level 2 items"""
        self.log("Analyzing hierarchy levels...")
        
        level1_items = [row for row in data_rows if row['_outline_level'] == 0]
        level2_items = [row for row in data_rows if row['_outline_level'] == 1]
        
        self.stats['level1_count'] = len(level1_items)
        self.stats['level2_count'] = len(level2_items)
        
        self.log(f"Level 1 items (outline_level=0): {len(level1_items)}")
        if level1_items:
            for item in level1_items[:3]:
                bom_line = item.get('BOM Line', 'N/A')
                self.log(f"  - Row {item['_row_num']}: {bom_line}")
        
        self.log(f"Level 2 items (outline_level=1): {len(level2_items)}")
        if level2_items:
            for item in level2_items[:5]:
                bom_line = item.get('BOM Line', 'N/A')
                self.log(f"  - Row {item['_row_num']}: {bom_line}")
        
        # Verify Level 1 has only one item
        if len(level1_items) != 1:
            self.log(f"WARNING: Expected 1 Level 1 item, found {len(level1_items)}", "WARNING")
        else:
            self.log("✓ Verified: Level 1 has exactly 1 item")
        
        return level1_items, level2_items
    
    def extract_level2_branches(self, data_rows: List[Dict], level2_items: List[Dict]) -> Dict[str, List[Dict]]:
        """Extract Level 2 items and their children into separate groups"""
        self.log("Extracting Level 2 branches...")
        
        branches = {}
        
        for i, level2_item in enumerate(level2_items):
            bom_line = level2_item.get('BOM Line', '')
            
            # Extract sheet name from text after the last '-' in BOM Line
            if bom_line and '-' in bom_line:
                # Find the last '-' and get everything after it
                sheet_name = bom_line.split('-')[-1].strip().upper()
                if not sheet_name:  # If nothing after last '-'
                    sheet_name = f"L2_{i+1}"
                    self.log(f"WARNING: Empty text after last '-' in BOM Line '{bom_line}', using {sheet_name}", "WARNING")
            else:
                sheet_name = f"L2_{i+1}"
                self.log(f"WARNING: No '-' found in BOM Line '{bom_line}', using {sheet_name}", "WARNING")

            self.log(f"  Extracted sheet name '{sheet_name}' from BOM Line: {bom_line}")

            # Handle duplicate sheet names
            original_name = sheet_name
            counter = 1
            while sheet_name in branches:
                sheet_name = f"{original_name}_{counter}"
                counter += 1
            
            # Find all children of this Level 2 item
            level2_row_num = level2_item['_row_num']
            
            # Determine the range of children (until next Level 2 or end)
            next_level2_row = None
            if i + 1 < len(level2_items):
                next_level2_row = level2_items[i + 1]['_row_num']
            
            # Collect Level 2 item and all its children
            branch_rows = [level2_item]
            
            for row in data_rows:
                row_num = row['_row_num']
                
                # Skip if before current Level 2 item
                if row_num <= level2_row_num:
                    continue
                
                # Stop if reached next Level 2 item
                if next_level2_row and row_num >= next_level2_row:
                    break
                
                # Add if it's a child (outline_level >= 2)
                if row['_outline_level'] >= 2:
                    branch_rows.append(row)
            
            branches[sheet_name] = branch_rows
            self.log(f"Branch '{sheet_name}': {len(branch_rows)} rows (Level 2 + children)")
            self.stats['rows_per_sheet'][sheet_name] = len(branch_rows)
        
        return branches
    
    def create_output_workbook(self, headers: Dict[str, int], branches: Dict[str, List[Dict]]) -> Any:
        """Create output workbook with extracted sheets"""
        self.log("Creating output workbook...")
        
        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)  # Remove default sheet
        
        # Sort headers by column index
        sorted_headers = sorted(headers.items(), key=lambda x: x[1])
        header_names = [h[0] for h in sorted_headers]
        
        for sheet_name, rows in branches.items():
            self.log(f"Creating sheet '{sheet_name}' with {len(rows)} rows")
            
            ws = output_wb.create_sheet(title=sheet_name)
            
            # Write headers
            for col_idx, header_name in enumerate(header_names, start=1):
                ws.cell(1, col_idx, header_name)
            
            # Write data rows
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, header_name in enumerate(header_names, start=1):
                    value = row_data.get(header_name)
                    ws.cell(row_idx, col_idx, value)
            
            self.stats['sheets_created'].append(sheet_name)
        
        self.log(f"Created {len(branches)} extraction sheets")
        return output_wb
    
    def create_compressed_sheets(self, output_wb: Any, branches: Dict[str, List[Dict]]):
        """Create compressed sheets with deduplication and quantity aggregation"""
        self.log("Creating compressed sheets...")
        
        # 🔧 使用配置获取输出列
        compress_columns = self.CONFIG.COMPRESS_OUTPUT_COLUMNS
        
        for sheet_name, rows in branches.items():
            compress_sheet_name = f"{sheet_name}_Compress"
            self.log(f"Creating compressed sheet '{compress_sheet_name}'")
            
            ws = output_wb.create_sheet(title=compress_sheet_name)
            
            # Write headers
            for col_idx, col_name in enumerate(compress_columns, start=1):
                ws.cell(1, col_idx, col_name)
            
            # 🔧 使用配置初始化分组数据
            grouped_data = defaultdict(lambda: self.CONFIG.get_metadata_dict())
            
            original_count = 0
            for row_data in rows:
                # 🔧 使用配置获取分组键
                grouping_keys = tuple(row_data.get(col) for col in self.CONFIG.GROUPING_KEY_COLUMNS)
                
                # Skip rows without part number
                if not grouping_keys[0]:
                    continue
                
                original_count += 1
                
                # Get and normalize quantity
                qty = row_data.get(self.CONFIG.ACCUMULATION_COLUMN)
                if qty is None or qty == '' or qty == 0 or qty == '0' or qty == 0.0:
                    qty = self.CONFIG.DEFAULT_QUANTITY
                else:
                    try:
                        qty = float(qty)
                    except (ValueError, TypeError):
                        self.log(f"WARNING: Invalid quantity '{qty}' in {sheet_name}, treating as {self.CONFIG.DEFAULT_QUANTITY}", "WARNING")
                        qty = self.CONFIG.DEFAULT_QUANTITY
                
                # 🔧 使用配置存储元数据
                if grouped_data[grouping_keys][self.CONFIG.METADATA_COLUMNS[0]] is None:
                    for col in self.CONFIG.METADATA_COLUMNS:
                        grouped_data[grouping_keys][col] = row_data.get(col)
                
                # Accumulate quantity
                grouped_data[grouping_keys][self.CONFIG.ACCUMULATION_COLUMN] += qty
            
            # 🔧 使用配置写入数据
            col_indices = self.CONFIG.get_compress_column_indices()
            row_idx = 2
            for keys, data in sorted(grouped_data.items()):
                for col_name in compress_columns:
                    col_idx = col_indices[col_name]
                    if col_name in self.CONFIG.GROUPING_KEY_COLUMNS:
                        # 写入分组键
                        key_idx = self.CONFIG.GROUPING_KEY_COLUMNS.index(col_name)
                        ws.cell(row_idx, col_idx, keys[key_idx])
                    else:
                        # 写入数据
                        ws.cell(row_idx, col_idx, data.get(col_name))
                row_idx += 1
            
            deduplicated_count = len(grouped_data)
            removed_count = original_count - deduplicated_count
            
            self.stats['dedup_stats'][compress_sheet_name] = {
                'original': original_count,
                'deduplicated': deduplicated_count,
                'removed': removed_count
            }
            
            self.log(f"  Original rows: {original_count}, After dedup: {deduplicated_count}, Removed: {removed_count}")
            self.stats['sheets_created'].append(compress_sheet_name)
            self.stats['rows_per_sheet'][compress_sheet_name] = deduplicated_count
        
        self.log(f"Created {len(branches)} compressed sheets")
    
    def create_unified_summary(self, output_wb: Any) -> openpyxl.Workbook:
        """Create unified BOM summary sheet consolidating all compress sheets."""
        self.log("Creating unified BOM summary...")
        
        # Collect compress sheets
        compress_sheets = {}
        for sheet_name in output_wb.sheetnames:
            if sheet_name.endswith('_Compress'):
                base_name = sheet_name.replace('_Compress', '')
                compress_sheets[base_name] = output_wb[sheet_name]
        
        if not compress_sheets:
            self.log("WARNING: No compress sheets found to create summary", "WARNING")
            return
        
        self.log(f"Found {len(compress_sheets)} compress sheets to consolidate")
        
        # 🔧 使用配置初始化统一数据
        unified_data = defaultdict(lambda: {
            **{col: None for col in self.CONFIG.METADATA_COLUMNS},
            'quantities': defaultdict(float),
            'metadata_sources': set()
        })
        
        # 🔧 使用配置获取列索引
        col_indices = self.CONFIG.get_compress_column_indices()
        
        # Collect data from compress sheets
        for sheet_type, ws in compress_sheets.items():
            self.log(f"  Processing compress sheet: {sheet_type}")
            
            for row_idx in range(2, ws.max_row + 1):
                # 🔧 动态读取分组键
                grouping_keys = tuple(
                    ws.cell(row_idx, col_indices[col]).value 
                    for col in self.CONFIG.GROUPING_KEY_COLUMNS
                )
                
                if not grouping_keys[0]:
                    continue
                
                # 🔧 动态读取元数据
                if unified_data[grouping_keys][self.CONFIG.METADATA_COLUMNS[0]] is None:
                    for col in self.CONFIG.METADATA_COLUMNS:
                        unified_data[grouping_keys][col] = ws.cell(row_idx, col_indices[col]).value
                    unified_data[grouping_keys]['metadata_sources'].add(sheet_type)
                
                # 读取数量
                quantity = ws.cell(row_idx, col_indices[self.CONFIG.ACCUMULATION_COLUMN]).value
                unified_data[grouping_keys]['quantities'][sheet_type] = quantity if quantity else 0
        
        # Create summary workbook
        summary_wb = openpyxl.Workbook()
        summary_ws = summary_wb.active
        summary_ws.title = 'BOM_Summary'
        
        sheet_types = list(compress_sheets.keys())
        
        # 🔧 使用配置定义表头
        headers = list(self.CONFIG.SUMMARY_FIXED_COLUMNS)
        headers.extend(sheet_types)
        
        for col_idx, header in enumerate(headers, start=1):
            summary_ws.cell(1, col_idx, header)
        
        # Write data
        row_idx = 2
        for keys, data in sorted(unified_data.items()):
            col_idx = 1
            
            # 写入分组键
            for key in keys:
                summary_ws.cell(row_idx, col_idx, key)
                col_idx += 1
            
            # 🔧 写入元数据
            for col_name in self.CONFIG.METADATA_COLUMNS:
                summary_ws.cell(row_idx, col_idx, data[col_name])
                col_idx += 1
            
            # 写入各sheet的数量
            for sheet_type in sheet_types:
                qty = data['quantities'].get(sheet_type, 0)
                summary_ws.cell(row_idx, col_idx, qty)
                col_idx += 1
            
            row_idx += 1
        
        unique_parts = len(unified_data)
        self.log(f"✓ Created BOM_Summary with {unique_parts} unique parts across {len(sheet_types)} sheet types")
        self.stats['sheets_created'].append('BOM_Summary')
        self.stats['rows_per_sheet']['BOM_Summary'] = unique_parts
        
        return summary_wb

    def process_from_buffer(self, uploaded_file_buffer, filename="uploaded_file.xlsm"):
        """
        Process BOM from uploaded file buffer (for Web interface)
        
        Args:
            uploaded_file_buffer: File-like object (BytesIO) from file upload
            filename: Original filename for logging
            
        Returns:
            tuple: (output_buffer, summary_buffer, log_text, stats_dict)
                - output_buffer: BytesIO containing extraction workbook
                - summary_buffer: BytesIO containing BOM summary
                - log_text: Complete log as text string
                - stats_dict: Dictionary with processing statistics
        """
        from io import BytesIO
        
        try:
            self.log("="*80)
            self.log("BOM PROCESSING STARTED (Web Mode)")
            self.log("="*80)
            self.log(f"Processing uploaded file: {filename}")
            
            # Temporarily set input_file for logging purposes
            self.input_file = filename
            
            # Step 1: Load workbook from buffer
            self.log(f"Loading workbook from uploaded file...")
            wb = load_workbook(uploaded_file_buffer, data_only=True, keep_vba=False)
            self.log(f"Workbook loaded successfully. Sheets: {wb.sheetnames}")
            
            # Get the main sheet (first sheet)
            main_sheet = wb.worksheets[0]
            sheet_name = main_sheet.title
            self.log(f"Processing main sheet: '{sheet_name}' ({main_sheet.max_row} rows, {main_sheet.max_column} cols)")
            
            # Extract headers from row 1
            headers = {}
            for col_idx in range(1, main_sheet.max_column + 1):
                cell_value = main_sheet.cell(1, col_idx).value
                if cell_value:
                    headers[str(cell_value).strip()] = col_idx
            
            self.log(f"Found {len(headers)} column headers")
            
            # Required columns validation
            required_cols = ['BOM Line', 'Weight', 'Area', 'CAD OEM Part Number', 'CAD OEM Rev', 'Quantity']
            missing_cols = [col for col in required_cols if col not in headers]
            if missing_cols:
                self.log(f"WARNING: Missing required columns: {missing_cols}", "WARNING")
            
            # Extract all data rows with outline levels
            data_rows = []
            for row_idx in range(2, main_sheet.max_row + 1):
                row_dim = main_sheet.row_dimensions[row_idx]
                outline_level = row_dim.outline_level if row_dim.outline_level else 0
                
                row_data = {'_row_num': row_idx, '_outline_level': outline_level}
                
                for col_name, col_idx in headers.items():
                    cell_value = main_sheet.cell(row_idx, col_idx).value
                    row_data[col_name] = cell_value
                
                data_rows.append(row_data)
            
            self.log(f"Extracted {len(data_rows)} data rows")
            
            # Step 2: Analyze hierarchy
            level1_items, level2_items = self.analyze_hierarchy(data_rows)
            
            # Step 3: Extract Level 2 branches
            branches = self.extract_level2_branches(data_rows, level2_items)
            
            # Step 4: Create output workbook with extracted sheets
            output_wb = self.create_output_workbook(headers, branches)
            
            # Step 5: Create compressed sheets
            self.create_compressed_sheets(output_wb, branches)
            
            # Step 6: Create unified BOM summary in separate file
            summary_wb = self.create_unified_summary(output_wb)
            
            # Step 7: Save to BytesIO buffers instead of files
            self.log("Preparing output files for download...")
            
            output_buffer = BytesIO()
            output_wb.save(output_buffer)
            output_buffer.seek(0)
            self.log("✓ Output workbook prepared")
            
            summary_buffer = BytesIO()
            summary_wb.save(summary_buffer)
            summary_buffer.seek(0)
            self.log("✓ BOM summary prepared")
            
            # Prepare log text
            log_text = self.get_log_content()
            self.log("✓ Log content prepared")
            
            self.log("="*80)
            self.log("BOM PROCESSING COMPLETED SUCCESSFULLY")
            self.log("="*80)
            
            return output_buffer, summary_buffer, log_text, self.stats
            
        except Exception as e:
            self.log(f"FATAL ERROR: {str(e)}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            raise
    
    def get_log_content(self):
        """Generate complete log content as string (for Web downloads)"""
        lines = []
        lines.append("="*80)
        lines.append("BOM PROCESSING LOG")
        lines.append("="*80)
        lines.append("")
        lines.append(f"Start Time: {self.start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Input File: {self.input_file}")
        lines.append("")
        lines.append("="*80)
        lines.append("PROCESSING LOG")
        lines.append("="*80)
        lines.append("")
        
        for msg in self.log_messages:
            lines.append(msg)
        
        lines.append("")
        lines.append("="*80)
        lines.append("STATISTICS SUMMARY")
        lines.append("="*80)
        lines.append("")
        lines.append(f"Level 1 items found: {self.stats['level1_count']}")
        lines.append(f"Level 2 items found: {self.stats['level2_count']}")
        lines.append(f"Sheets created: {len(self.stats['sheets_created'])}")
        lines.append(f"Sheet names: {', '.join(self.stats['sheets_created'])}")
        lines.append("")
        
        lines.append("Rows per sheet:")
        for sheet_name, count in self.stats['rows_per_sheet'].items():
            lines.append(f"  {sheet_name}: {count} rows")
        
        if self.stats['dedup_stats']:
            lines.append("")
            lines.append("Deduplication statistics:")
            for sheet_name, stats in self.stats['dedup_stats'].items():
                lines.append(f"  {sheet_name}:")
                lines.append(f"    Original rows: {stats['original']}")
                lines.append(f"    After dedup: {stats['deduplicated']}")
                lines.append(f"    Removed: {stats['removed']}")
        
        end_time = datetime.datetime.now()
        duration = end_time - self.start_time
        lines.append("")
        lines.append(f"End Time: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Total Duration: {duration.total_seconds():.2f} seconds")
        lines.append("")
        lines.append("="*80)
        lines.append("PROCESSING COMPLETED SUCCESSFULLY")
        lines.append("="*80)
        
        return "\n".join(lines)

    def process(self):
        """Main processing pipeline (for CLI usage)"""
        try:
            self.log("="*80)
            self.log("BOM PROCESSING STARTED")
            self.log("="*80)
            
            # Step 1: Load workbook
            wb, main_sheet, headers, data_rows = self.load_workbook_data()
            
            # Step 2: Analyze hierarchy
            level1_items, level2_items = self.analyze_hierarchy(data_rows)
            
            # Step 3: Extract Level 2 branches
            branches = self.extract_level2_branches(data_rows, level2_items)
            
            # Step 4: Create output workbook with extracted sheets
            output_wb = self.create_output_workbook(headers, branches)
            
            # Step 5: Create compressed sheets
            self.create_compressed_sheets(output_wb, branches)
            
            # Step 6: Create unified BOM summary in separate file
            summary_wb = self.create_unified_summary(output_wb)
            
            # Step 7: Save output workbook
            self.log(f"Saving output workbook to: {self.output_file}")
            output_wb.save(self.output_file)
            self.log("✓ Output workbook saved successfully")
            
            # Step 8: Save BOM summary workbook
            self.log(f"Saving BOM summary to: {self.summary_file}")
            summary_wb.save(self.summary_file)
            self.log("✓ BOM summary saved successfully")
            
            # Step 9: Write log file
            self.log(f"Writing log file to: {self.log_file}")
            self.write_log_file()
            
            self.log("="*80)
            self.log("BOM PROCESSING COMPLETED SUCCESSFULLY")
            self.log("="*80)
            
            return True
            
        except Exception as e:
            self.log(f"FATAL ERROR: {str(e)}", "ERROR")
            import traceback
            self.log(traceback.format_exc(), "ERROR")
            
            # Try to write log even on failure
            try:
                self.write_log_file()
            except:
                pass
            
            raise


def main():
    """Main entry point"""
    # Generate date stamp for output files
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    
    # File paths with date stamp
    input_file = r"c:\Users\zongyue.liu\Desktop\AdvSimulation\BOM_Summary\data\COS1000334701_BA-01302026_rev15.xlsm"
    file_name_wo_ext = os.path.splitext(os.path.basename(input_file))[0]
    current_time_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    output_file = rf"c:\Users\zongyue.liu\Desktop\AdvSimulation\BOM_Summary\data\{file_name_wo_ext}_processed_{current_time_str}.xlsx"
    summary_file = rf"c:\Users\zongyue.liu\Desktop\AdvSimulation\BOM_Summary\data\{file_name_wo_ext}_BOM_Sum_{current_time_str}.xlsx"
    log_file = rf"c:\Users\zongyue.liu\Desktop\AdvSimulation\BOM_Summary\data\{file_name_wo_ext}_process_log_{current_time_str}.txt"
    
    # Verify input file exists
    if not os.path.exists(input_file):
        print(f"ERROR: Input file not found: {input_file}")
        return
    
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print(f"Summary file: {summary_file}")
    print(f"Log file: {log_file}")
    print()
    
    # Create processor and run
    processor = BOMProcessor(input_file, output_file, summary_file, log_file)
    processor.process()
    
    print()
    print(f"✓ Processing complete!")
    print(f"  Output: {output_file}")
    print(f"  Summary: {summary_file}")
    print(f"  Log: {log_file}")


if __name__ == "__main__":
    main()

