"""
BOM Processor Regression Test Suite (优化版)
Author: Zongyue Liu
Date: 2026-01-29

完整的回归测试框架，验证BOM处理器的输出正确性：
- 批量提取后直接比对（大幅提升性能）
- 浮点数容差比较（默认1e-6）
- 统计数据验证
- 每个测试显示运行时间

使用方法：
    pytest tests/test_bom_processor.py -v
    pytest tests/test_bom_processor.py -v -s  # 显示详细输出
"""

import pytest
import openpyxl
from pathlib import Path
import sys
import json
import time
from typing import Dict, List, Tuple, Any, Optional
import functools

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from bom_processor import BOMProcessor


def timing_decorator(func):
    """装饰器：记录测试运行时间"""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        elapsed_time = time.time() - start_time
        print(f"\n⏱️  测试运行时间: {elapsed_time:.2f}秒")
        return result
    return wrapper


class BOMTestConfig:
    """测试配置"""
    # 浮点数比较容差
    FLOAT_TOLERANCE = 1e-6
    
    # 测试数据路径
    TEST_DATA_DIR = Path(__file__).parent / "test_data"
    BASELINE_DIR = TEST_DATA_DIR / "baseline"
    TEMP_OUTPUT_DIR = TEST_DATA_DIR / "temp_outputs"
    
    # 基准文件名
    INPUT_FILE = "BOM-COS1000334701-BA.xlsm"
    BASELINE_OUTPUT = "BOM-COS1000334701-BA_processed_baseline.xlsx"
    BASELINE_SUMMARY = "BOM-COS1000334701-BA_BOM_Sum_baseline.xlsx"
    BASELINE_STATS = "baseline_stats.json"
    
    # 性能阈值（秒）
    MAX_PROCESSING_TIME = 60


class CellComparisonResult:
    """单元格比较结果"""
    def __init__(self):
        self.differences: List[Dict] = []
        self.total_cells_compared = 0
        self.identical_cells = 0
        
    def add_difference(self, sheet: str, row: int, col: int, 
                      expected: Any, actual: Any, reason: str = ""):
        """记录差异"""
        self.differences.append({
            'sheet': sheet,
            'row': row,
            'col': col,
            'expected': expected,
            'actual': actual,
            'reason': reason
        })
    
    def record_match(self):
        """记录匹配"""
        self.identical_cells += 1
        self.total_cells_compared += 1
    
    def record_comparison(self):
        """记录比较"""
        self.total_cells_compared += 1
    
    @property
    def has_differences(self) -> bool:
        return len(self.differences) > 0
    
    @property
    def match_rate(self) -> float:
        if self.total_cells_compared == 0:
            return 0.0
        return (self.identical_cells / self.total_cells_compared) * 100
    
    def get_summary(self) -> str:
        """获取比较摘要"""
        summary = [
            f"单元格比较统计:",
            f"  总计: {self.total_cells_compared}",
            f"  匹配: {self.identical_cells}",
            f"  差异: {len(self.differences)}",
            f"  匹配率: {self.match_rate:.2f}%"
        ]
        
        if self.differences:
            summary.append(f"\n前10个差异:")
            for i, diff in enumerate(self.differences[:10], 1):
                summary.append(
                    f"  {i}. 工作表'{diff['sheet']}' "
                    f"行{diff['row']} 列{diff['col']}: "
                    f"期望={diff['expected']}, 实际={diff['actual']}"
                    f"{' (' + diff['reason'] + ')' if diff['reason'] else ''}"
                )
        
        return "\n".join(summary)


class BOMTestUtils:
    """测试工具类"""
    
    @staticmethod
    def compare_values(expected: Any, actual: Any, tolerance: float = BOMTestConfig.FLOAT_TOLERANCE) -> Tuple[bool, str]:
        """
        比较两个值，支持浮点数容差
        
        Returns:
            (is_equal, reason) - 是否相等及原因
        """
        # 都是None
        if expected is None and actual is None:
            return True, ""
        
        # 一个是None
        if expected is None or actual is None:
            return False, f"一个为None: expected={expected}, actual={actual}"
        
        # 尝试作为数字比较
        try:
            expected_float = float(expected)
            actual_float = float(actual)
            
            # 浮点数容差比较
            if abs(expected_float - actual_float) <= tolerance:
                return True, ""
            else:
                return False, f"数值差异超过容差{tolerance}: |{expected_float} - {actual_float}| = {abs(expected_float - actual_float)}"
        except (ValueError, TypeError):
            # 不是数字，作为字符串比较
            pass
        
        # 字符串比较（转换后比较）
        expected_str = str(expected).strip() if expected is not None else ""
        actual_str = str(actual).strip() if actual is not None else ""
        
        if expected_str == actual_str:
            return True, ""
        else:
            return False, f"字符串不匹配: '{expected_str}' != '{actual_str}'"
    
    @staticmethod
    def extract_worksheet_data(ws) -> Dict[Tuple[int, int], Any]:
        """
        一次性提取工作表所有数据到字典
        
        Args:
            ws: openpyxl worksheet
            
        Returns:
            字典 {(row, col): value}
        """
        data = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                data[(cell.row, cell.column)] = cell.value
        return data
    
    @staticmethod
    def compare_worksheets_fast(expected_ws, actual_ws, sheet_name: str) -> CellComparisonResult:
        """
        快速比较两个工作表（批量提取后直接比对）
        
        Args:
            expected_ws: 期望的工作表
            actual_ws: 实际的工作表
            sheet_name: 工作表名称（用于错误报告）
            
        Returns:
            CellComparisonResult对象
        """
        result = CellComparisonResult()
        
        # 比较维度
        if expected_ws.max_row != actual_ws.max_row:
            result.add_difference(
                sheet_name, 0, 0,
                f"行数={expected_ws.max_row}",
                f"行数={actual_ws.max_row}",
                "工作表行数不匹配"
            )
        
        if expected_ws.max_column != actual_ws.max_column:
            result.add_difference(
                sheet_name, 0, 0,
                f"列数={expected_ws.max_column}",
                f"列数={actual_ws.max_column}",
                "工作表列数不匹配"
            )
        
        # 一次性提取所有数据
        print(f"    提取期望数据...")
        expected_data = BOMTestUtils.extract_worksheet_data(expected_ws)
        print(f"    提取实际数据...")
        actual_data = BOMTestUtils.extract_worksheet_data(actual_ws)
        
        # 获取所有可能的单元格位置
        all_positions = set(expected_data.keys()) | set(actual_data.keys())
        total_cells = len(all_positions)
        
        print(f"    比较 {total_cells} 个单元格...")
        
        # 直接比对字典
        for position in all_positions:
            result.record_comparison()
            
            expected_value = expected_data.get(position)
            actual_value = actual_data.get(position)
            
            is_equal, reason = BOMTestUtils.compare_values(expected_value, actual_value)
            
            if is_equal:
                result.record_match()
            else:
                row, col = position
                result.add_difference(
                    sheet_name, row, col,
                    expected_value, actual_value,
                    reason
                )
        
        return result
    
    @staticmethod
    def compare_workbooks(expected_path: Path, actual_path: Path) -> CellComparisonResult:
        """
        比较两个工作簿的所有工作表（优化版）
        
        Args:
            expected_path: 期望的工作簿路径
            actual_path: 实际的工作簿路径
            
        Returns:
            CellComparisonResult对象（合并所有工作表的结果）
        """
        print(f"\n📊 比较工作簿:")
        print(f"  期望: {expected_path.name}")
        print(f"  实际: {actual_path.name}")
        
        start_load = time.time()
        expected_wb = openpyxl.load_workbook(expected_path, data_only=True)
        actual_wb = openpyxl.load_workbook(actual_path, data_only=True)
        load_time = time.time() - start_load
        print(f"  ⏱️  加载时间: {load_time:.2f}秒")
        
        combined_result = CellComparisonResult()
        
        # 比较工作表列表
        expected_sheets = set(expected_wb.sheetnames)
        actual_sheets = set(actual_wb.sheetnames)
        
        if expected_sheets != actual_sheets:
            missing_in_actual = expected_sheets - actual_sheets
            extra_in_actual = actual_sheets - expected_sheets
            
            if missing_in_actual:
                combined_result.add_difference(
                    "工作簿", 0, 0,
                    f"工作表: {sorted(expected_sheets)}",
                    f"缺少工作表: {sorted(missing_in_actual)}",
                    "工作表缺失"
                )
            
            if extra_in_actual:
                combined_result.add_difference(
                    "工作簿", 0, 0,
                    f"工作表: {sorted(expected_sheets)}",
                    f"额外工作表: {sorted(extra_in_actual)}",
                    "工作表多余"
                )
        
        # 比较每个工作表
        common_sheets = expected_sheets & actual_sheets
        for i, sheet_name in enumerate(sorted(common_sheets), 1):
            print(f"\n  [{i}/{len(common_sheets)}] 工作表: {sheet_name}")
            expected_ws = expected_wb[sheet_name]
            actual_ws = actual_wb[sheet_name]
            
            sheet_start = time.time()
            sheet_result = BOMTestUtils.compare_worksheets_fast(expected_ws, actual_ws, sheet_name)
            sheet_time = time.time() - sheet_start
            
            # 合并结果
            combined_result.differences.extend(sheet_result.differences)
            combined_result.total_cells_compared += sheet_result.total_cells_compared
            combined_result.identical_cells += sheet_result.identical_cells
            
            status = "✓" if not sheet_result.has_differences else "✗"
            print(f"    {status} 单元格: {sheet_result.total_cells_compared}, "
                  f"匹配: {sheet_result.identical_cells}, "
                  f"差异: {len(sheet_result.differences)}, "
                  f"耗时: {sheet_time:.2f}秒")
        
        return combined_result


class TestBOMProcessor:
    """BOM处理器测试类"""
    
    @pytest.fixture(scope="class", autouse=True)
    def setup_test_environment(self):
        """设置测试环境"""
        # 确保临时输出目录存在且为空
        BOMTestConfig.TEMP_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        # 清理之前的临时文件
        for file in BOMTestConfig.TEMP_OUTPUT_DIR.glob("*"):
            if file.is_file():
                try:
                    file.unlink()
                except:
                    pass
        
        yield
        
        # 测试后不删除临时文件，便于调试
        print(f"\n💾 临时文件保留在: {BOMTestConfig.TEMP_OUTPUT_DIR}")
    
    @timing_decorator
    def test_baseline_files_exist(self):
        """测试1: 验证基准文件存在"""
        input_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.INPUT_FILE
        
        if not input_file.exists():
            pytest.skip(
                f"基准输入文件不存在: {input_file}\n"
                f"请将 'BOM-COS1000334701-BA.xlsm' 复制到 {BOMTestConfig.BASELINE_DIR}"
            )
        
        # 如果没有基准输出文件，生成它们
        baseline_output = BOMTestConfig.BASELINE_DIR / BOMTestConfig.BASELINE_OUTPUT
        baseline_summary = BOMTestConfig.BASELINE_DIR / BOMTestConfig.BASELINE_SUMMARY
        baseline_stats = BOMTestConfig.BASELINE_DIR / BOMTestConfig.BASELINE_STATS
        
        if not baseline_output.exists() or not baseline_summary.exists():
            print("\n📝 生成基准输出文件...")
            
            processor = BOMProcessor(
                str(input_file),
                str(baseline_output),
                str(baseline_summary),
                str(BOMTestConfig.BASELINE_DIR / "baseline_log.txt")
            )
            
            success = processor.process()
            assert success, "基准文件生成失败"
            
            # 保存统计信息
            baseline_stats.write_text(
                json.dumps(processor.stats, indent=2, ensure_ascii=False),
                encoding='utf-8'
            )
            
            print(f"✓ 基准文件已生成:")
            print(f"  - {baseline_output.name}")
            print(f"  - {baseline_summary.name}")
            print(f"  - {baseline_stats.name}")
            pytest.skip("首次运行：已生成基准文件，请再次运行测试进行验证")
        
        assert baseline_output.exists(), f"基准输出文件不存在: {baseline_output}"
        assert baseline_summary.exists(), f"基准汇总文件不存在: {baseline_summary}"
        assert baseline_stats.exists(), f"基准统计文件不存在: {baseline_stats}"
    
    @timing_decorator
    def test_processor_execution(self):
        """测试2: 验证处理器成功执行"""
        input_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.INPUT_FILE
        
        if not input_file.exists():
            pytest.skip(f"输入文件不存在: {input_file}")
        
        output_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_output.xlsx"
        summary_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_summary.xlsx"
        log_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_log.txt"
        
        processor = BOMProcessor(
            str(input_file),
            str(output_file),
            str(summary_file),
            str(log_file)
        )
        
        start_time = time.time()
        success = processor.process()
        elapsed_time = time.time() - start_time
        
        assert success, "BOM处理失败"
        assert output_file.exists(), "输出文件未生成"
        assert summary_file.exists(), "汇总文件未生成"
        assert log_file.exists(), "日志文件未生成"
        assert elapsed_time < BOMTestConfig.MAX_PROCESSING_TIME, \
            f"处理时间超时: {elapsed_time:.1f}秒 > {BOMTestConfig.MAX_PROCESSING_TIME}秒"
        
        print(f"\n✓ 处理成功 (耗时: {elapsed_time:.2f}秒)")
    
    @timing_decorator
    def test_statistics_validation(self):
        """测试3: 验证统计数据一致性"""
        baseline_stats_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.BASELINE_STATS
        
        if not baseline_stats_file.exists():
            pytest.skip("基准统计文件不存在")
        
        # 运行处理器
        input_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.INPUT_FILE
        output_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_output.xlsx"
        summary_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_summary.xlsx"
        log_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_log.txt"
        
        processor = BOMProcessor(
            str(input_file),
            str(output_file),
            str(summary_file),
            str(log_file)
        )
        processor.process()
        
        # 加载基准统计
        baseline_stats = json.loads(baseline_stats_file.read_text(encoding='utf-8'))
        actual_stats = processor.stats
        
        # 验证关键统计
        print("\n📊 统计对比:")
        print(f"  Level 1: 基准={baseline_stats['level1_count']}, 实际={actual_stats['level1_count']}")
        print(f"  Level 2: 基准={baseline_stats['level2_count']}, 实际={actual_stats['level2_count']}")
        print(f"  创建的工作表数: 基准={len(baseline_stats['sheets_created'])}, 实际={len(actual_stats['sheets_created'])}")
        
        assert actual_stats['level1_count'] == baseline_stats['level1_count'], \
            f"Level 1 项目数不匹配: {actual_stats['level1_count']} != {baseline_stats['level1_count']}"
        
        assert actual_stats['level2_count'] == baseline_stats['level2_count'], \
            f"Level 2 项目数不匹配: {actual_stats['level2_count']} != {baseline_stats['level2_count']}"
        
        assert set(actual_stats['sheets_created']) == set(baseline_stats['sheets_created']), \
            f"创建的工作表不匹配:\n基准: {sorted(baseline_stats['sheets_created'])}\n实际: {sorted(actual_stats['sheets_created'])}"
        
        # 验证每个工作表的行数
        for sheet_name, baseline_rows in baseline_stats['rows_per_sheet'].items():
            actual_rows = actual_stats['rows_per_sheet'].get(sheet_name)
            assert actual_rows == baseline_rows, \
                f"工作表 '{sheet_name}' 行数不匹配: {actual_rows} != {baseline_rows}"
        
        # 验证去重统计
        if baseline_stats.get('dedup_stats'):
            for sheet_name, baseline_dedup in baseline_stats['dedup_stats'].items():
                actual_dedup = actual_stats['dedup_stats'].get(sheet_name, {})
                assert actual_dedup.get('original') == baseline_dedup.get('original'), \
                    f"工作表 '{sheet_name}' 原始行数不匹配"
                assert actual_dedup.get('deduplicated') == baseline_dedup.get('deduplicated'), \
                    f"工作表 '{sheet_name}' 去重后行数不匹配"
        
        print("✓ 所有统计数据匹配")
    
    @timing_decorator
    def test_output_file_cell_by_cell_comparison(self):
        """测试4: 输出文件逐单元格比对（优化版）"""
        baseline_output = BOMTestConfig.BASELINE_DIR / BOMTestConfig.BASELINE_OUTPUT
        actual_output = BOMTestConfig.TEMP_OUTPUT_DIR / "test_output.xlsx"
        
        if not baseline_output.exists():
            pytest.skip("基准输出文件不存在")
        
        if not actual_output.exists():
            # 如果测试输出文件不存在，先运行处理器
            input_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.INPUT_FILE
            summary_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_summary.xlsx"
            log_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_log.txt"
            
            processor = BOMProcessor(
                str(input_file),
                str(actual_output),
                str(summary_file),
                str(log_file)
            )
            processor.process()
        
        # 逐单元格比较（批量提取优化）
        result = BOMTestUtils.compare_workbooks(baseline_output, actual_output)
        
        print(f"\n{result.get_summary()}")
        
        # 断言：不应有差异
        if result.has_differences:
            # 保存差异报告
            diff_report_path = BOMTestConfig.TEMP_OUTPUT_DIR / "differences_output.json"
            diff_report_path.write_text(
                json.dumps(result.differences, indent=2, ensure_ascii=False),
                encoding='utf-8'
            )
            print(f"\n📄 完整差异报告已保存至: {diff_report_path}")
        
        assert not result.has_differences, \
            f"输出文件存在 {len(result.differences)} 个差异，匹配率: {result.match_rate:.2f}%"
        
        print("✓ 输出文件完全匹配（逐单元格）")
    
    @timing_decorator
    def test_summary_file_cell_by_cell_comparison(self):
        """测试5: BOM汇总文件逐单元格比对（优化版）"""
        baseline_summary = BOMTestConfig.BASELINE_DIR / BOMTestConfig.BASELINE_SUMMARY
        actual_summary = BOMTestConfig.TEMP_OUTPUT_DIR / "test_summary.xlsx"
        
        if not baseline_summary.exists():
            pytest.skip("基准汇总文件不存在")
        
        if not actual_summary.exists():
            # 如果测试汇总文件不存在，先运行处理器
            input_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.INPUT_FILE
            output_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_output.xlsx"
            log_file = BOMTestConfig.TEMP_OUTPUT_DIR / "test_log.txt"
            
            processor = BOMProcessor(
                str(input_file),
                str(output_file),
                str(actual_summary),
                str(log_file)
            )
            processor.process()
        
        # 逐单元格比较（批量提取优化）
        result = BOMTestUtils.compare_workbooks(baseline_summary, actual_summary)
        
        print(f"\n{result.get_summary()}")
        
        # 断言：不应有差异
        if result.has_differences:
            # 保存差异报告
            diff_report_path = BOMTestConfig.TEMP_OUTPUT_DIR / "differences_summary.json"
            diff_report_path.write_text(
                json.dumps(result.differences, indent=2, ensure_ascii=False),
                encoding='utf-8'
            )
            print(f"\n📄 完整差异报告已保存至: {diff_report_path}")
        
        assert not result.has_differences, \
            f"汇总文件存在 {len(result.differences)} 个差异，匹配率: {result.match_rate:.2f}%"
        
        print("✓ BOM汇总文件完全匹配（逐单元格）")
    
    @timing_decorator
    def test_processing_performance(self):
        """测试6: 处理性能验证（单次运行）"""
        input_file = BOMTestConfig.BASELINE_DIR / BOMTestConfig.INPUT_FILE
        
        if not input_file.exists():
            pytest.skip("输入文件不存在")
        
        output_file = BOMTestConfig.TEMP_OUTPUT_DIR / "perf_test_output.xlsx"
        summary_file = BOMTestConfig.TEMP_OUTPUT_DIR / "perf_test_summary.xlsx"
        log_file = BOMTestConfig.TEMP_OUTPUT_DIR / "perf_test_log.txt"
        
        processor = BOMProcessor(
            str(input_file),
            str(output_file),
            str(summary_file),
            str(log_file)
        )
        
        # 只运行1次（不重复3次，节省时间）
        start_time = time.time()
        processor.process()
        elapsed_time = time.time() - start_time
        
        print(f"\n⏱️  性能统计:")
        print(f"  处理时间: {elapsed_time:.2f}秒")
        print(f"  阈值: {BOMTestConfig.MAX_PROCESSING_TIME}秒")
        
        assert elapsed_time < BOMTestConfig.MAX_PROCESSING_TIME, \
            f"处理时间超过阈值: {elapsed_time:.2f}秒 > {BOMTestConfig.MAX_PROCESSING_TIME}秒"
        
        print(f"✓ 性能达标")


if __name__ == "__main__":
    # 直接运行此文件时使用pytest
    pytest.main([__file__, "-v", "-s"])
