"""
BOM Processor - Streamlit Web Application
Author: Zongyue Liu
Date: 2026-01-27

Streamlit web interface for BOM hierarchical processing
"""

import streamlit as st
import sys
import os
from pathlib import Path
from datetime import datetime
from io import BytesIO
import time

# Add parent directory to path to import bom_processor
sys.path.append(str(Path(__file__).parent.parent))

from bom_processor import BOMProcessor
import config
from utils.cleanup import cleanup_old_temp_files, get_temp_dir_info


# Page configuration
st.set_page_config(
    page_title=config.APP_TITLE,
    page_icon=config.APP_ICON,
    layout=config.PAGE_LAYOUT
)


def initialize_session_state():
    """Initialize session state variables"""
    if 'initialized' not in st.session_state:
        st.session_state.initialized = True
        st.session_state.processing = False
        st.session_state.processed = False
        st.session_state.output_buffer = None
        st.session_state.summary_buffer = None
        st.session_state.log_text = None
        st.session_state.stats = None
        st.session_state.uploaded_filename = None
        st.session_state.log_messages = []
        
        # Run cleanup on first initialization
        deleted, errors = cleanup_old_temp_files(config.TEMP_DIR, config.CLEANUP_HOURS)
        if deleted > 0:
            print(f"Startup cleanup: Deleted {deleted} old temp directories")


def validate_excel_file(uploaded_file):
    """
    Validate uploaded Excel file structure
    
    Returns:
        tuple: (is_valid, missing_columns, warnings)
    """
    try:
        from openpyxl import load_workbook
        
        # Load workbook to check structure
        wb = load_workbook(uploaded_file, data_only=True, keep_vba=False)
        
        if not wb.worksheets:
            return False, [], ["No worksheets found in file"]
        
        # Get first sheet
        sheet = wb.worksheets[0]
        
        # Extract headers
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(1, col_idx).value
            if cell_value:
                headers.append(str(cell_value).strip())
        
        # Check required columns
        missing_cols = [col for col in config.REQUIRED_COLUMNS if col not in headers]
        
        warnings = []
        if missing_cols:
            warnings.append(f"Missing required columns: {', '.join(missing_cols)}")
        
        # Check if file has data rows
        if sheet.max_row < 2:
            warnings.append("File has no data rows (only headers)")
        
        # Reset file pointer
        uploaded_file.seek(0)
        
        return len(warnings) == 0, missing_cols, warnings
        
    except Exception as e:
        uploaded_file.seek(0)
        return False, [], [f"File validation error: {str(e)}"]


def log_callback(log_entry):
    """Callback function to receive log messages from BOMProcessor"""
    st.session_state.log_messages.append(log_entry)


def main():
    """Main application"""
    
    # Initialize session state
    initialize_session_state()
    
    # Header
    st.title(f"{config.APP_ICON} {config.APP_TITLE}")
    st.markdown("---")
    
    # Create layout: left column for upload, right column for logs
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.header("📁 File Upload")
        
        # File uploader
        uploaded_file = st.file_uploader(
            "Upload BOM Excel File",
            type=['xlsm', 'xlsx'],
            help=f"Supports .xlsm and .xlsx files, max {config.MAX_UPLOAD_SIZE_MB}MB"
        )
        
        if uploaded_file is not None:
            st.session_state.uploaded_filename = uploaded_file.name
            
            # Show file info
            file_size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
            st.success(f"✅ Uploaded: {uploaded_file.name}")
            st.caption(f"File size: {file_size_mb:.2f} MB")
            
            # Validate file
            with st.spinner("Validating file structure..."):
                is_valid, missing_cols, warnings = validate_excel_file(uploaded_file)
            
            if warnings:
                st.warning("⚠️ File Validation Warnings:")
                for warning in warnings:
                    st.markdown(f"- {warning}")
                st.info("Warning: Processing will continue, but may fail. Please confirm the file format is correct.")
            
            st.markdown("---")
            
            # Process button
            process_button = st.button(
                "🚀 Start Processing",
                type="primary",
                disabled=st.session_state.processing,
                use_container_width=True
            )
            
            if process_button:
                st.session_state.processing = True
                st.session_state.processed = False
                st.session_state.log_messages = []
                st.rerun()
        
        else:
            st.info("👆 Please upload a BOM Excel file to start processing")
        
        # Temp directory info (collapsed by default)
        with st.expander("🗂️ Temporary Files Info"):
            temp_info = get_temp_dir_info(config.TEMP_DIR)
            st.metric("Temp Folders", temp_info['folder_count'])
            st.metric("Space Used", f"{temp_info['total_size_mb']:.2f} MB")
            if temp_info['oldest_folder']:
                st.caption(f"Oldest folder: {temp_info['oldest_folder']}")
    
    with col2:
        st.header("📋 Processing Log")
        
        # Log display area
        log_container = st.empty()
        
        # Process the file if processing flag is set
        if st.session_state.processing:
            
            try:
                # Create processor with log callback
                processor = BOMProcessor(log_callback=log_callback)
                
                # Reset file pointer
                uploaded_file.seek(0)
                
                # Process with spinner for each major step
                with st.spinner("Processing BOM file..."):
                    # Start processing
                    output_buffer, summary_buffer, log_text, stats = processor.process_from_buffer(
                        uploaded_file,
                        st.session_state.uploaded_filename
                    )
                
                # Store results in session state
                st.session_state.output_buffer = output_buffer
                st.session_state.summary_buffer = summary_buffer
                st.session_state.log_text = log_text
                st.session_state.stats = stats
                st.session_state.processed = True
                st.session_state.processing = False
                
                st.rerun()
                
            except Exception as e:
                st.session_state.processing = False
                st.error(f"❌ Processing failed: {str(e)}")
                
                # Show error log
                if st.session_state.log_messages:
                    st.text_area(
                        "Error Log",
                        "\n".join(st.session_state.log_messages),
                        height=300
                    )
                
                # Offer to download error log
                if st.session_state.log_messages:
                    error_log = "\n".join(st.session_state.log_messages)
                    st.download_button(
                        "📥 Download Error Log",
                        error_log,
                        file_name=f"error_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                        mime="text/plain"
                    )
        
        # Display logs
        if st.session_state.log_messages:
            log_text_display = "\n".join(st.session_state.log_messages)
            log_container.text_area(
                "Real-time Log",
                log_text_display,
                height=500,
                key=f"log_display_{len(st.session_state.log_messages)}"
            )
        else:
            log_container.info("📝 Processing logs will be displayed here...")
    
    # Results section (below the columns)
    if st.session_state.processed:
        st.markdown("---")
        st.header("✅ Processing Complete")
        
        # Statistics summary
        stats = st.session_state.stats
        
        col_s1, col_s2, col_s3, col_s4 = st.columns(4)
        with col_s1:
            st.metric("Level 1 Items", stats['level1_count'])
        with col_s2:
            st.metric("Level 2 Items", stats['level2_count'])
        with col_s3:
            st.metric("Sheets Created", len(stats['sheets_created']))
        with col_s4:
            if 'BOM_Summary' in stats['rows_per_sheet']:
                st.metric("Unique Parts", stats['rows_per_sheet']['BOM_Summary'])
        
        st.markdown("---")
        st.subheader("📥 Download Result Files")
        
        # Download buttons in three columns
        col_d1, col_d2, col_d3 = st.columns(3)
        
        current_date = datetime.now().strftime("%Y%m%d")
        base_filename = Path(st.session_state.uploaded_filename).stem
        
        with col_d1:
            st.download_button(
                "📊 Download Extraction Workbook",
                st.session_state.output_buffer.getvalue(),
                file_name=f"{base_filename}_Extracted_{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.caption("Contains all Level 2 branches and compressed sheets")
        
        with col_d2:
            st.download_button(
                "📈 Download BOM Summary",
                st.session_state.summary_buffer.getvalue(),
                file_name=f"BOM_Summary_{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.caption("Contains consolidated parts summary")
        
        with col_d3:
            st.download_button(
                "📄 Download Processing Log",
                st.session_state.log_text,
                file_name=f"processing_log_{current_date}.txt",
                mime="text/plain",
                use_container_width=True
            )
            st.caption("Contains detailed processing steps and statistics")
        
        st.markdown("---")
        
        # Reset button
        if st.button("🔄 Process New File", use_container_width=True):
            # Clear session state
            st.session_state.processing = False
            st.session_state.processed = False
            st.session_state.output_buffer = None
            st.session_state.summary_buffer = None
            st.session_state.log_text = None
            st.session_state.stats = None
            st.session_state.uploaded_filename = None
            st.session_state.log_messages = []
            st.rerun()
    
    # Footer
    st.markdown("---")
    st.caption(f"BOM Processor Web Application | Version 1.0 | {datetime.now().year}")


if __name__ == "__main__":
    main()
