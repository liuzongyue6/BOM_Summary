"""
BOM Comparison Utility
Author: Zongyue Liu
Date: 2026-01-27

Utility for comparing two BOM Excel files and generating comparison reports
"""

import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
from io import BytesIO
import sys

# Add parent directory to path
sys.path.append(str(Path(__file__).parent.parent))
import config


class BOMComparator:
    """Compare two BOM Excel files and generate comparison report"""
    
    def __init__(self, file1_bytes: bytes, file1_name: str, file2_bytes: bytes, file2_name: str):
        """
        Initialize BOM Comparator
        
        Args:
            file1_bytes: Bytes content of first Excel file
            file1_name: Filename of first Excel file
            file2_bytes: Bytes content of second Excel file
            file2_name: Filename of second Excel file
        """
        self.file1_name = Path(file1_name).stem
        self.file2_name = Path(file2_name).stem
        
        # Load workbooks
        self.wb1 = openpyxl.load_workbook(BytesIO(file1_bytes), data_only=True)
        self.wb2 = openpyxl.load_workbook(BytesIO(file2_bytes), data_only=True)
        
        # Get first worksheets
        self.ws1 = self.wb1.worksheets[0]
        self.ws2 = self.wb2.worksheets[0]
        
        # Data storage
        self.data1: Dict[Tuple, Dict] = {}
        self.data2: Dict[Tuple, Dict] = {}
        
        # Comparison results
        self.matching_keys: List[Tuple] = []
        self.unique_file1: List[Tuple] = []
        self.unique_file2: List[Tuple] = []
        
        # Yellow fill for differences
        self.yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    def sanitize_sheet_name(self, name: str, max_length: int = 31) -> str:
        """
        Sanitize sheet name to comply with Excel requirements
        
        Args:
            name: Sheet name to sanitize
            max_length: Maximum length (Excel limit is 31)
        
        Returns:
            Sanitized sheet name
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')
        
        # Truncate to max length
        if len(name) > max_length:
            name = name[:max_length]
        
        return name
    
    def extract_headers(self, worksheet: Any) -> Dict[str, int]:
        """
        Extract column headers from worksheet
        
        Args:
            worksheet: openpyxl worksheet
        
        Returns:
            Dictionary mapping column names to column indices
        """
        headers = {}
        for col_idx, cell in enumerate(worksheet[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx
        return headers
    
    def validate_columns(self, headers: Dict[str, int], filename: str) -> Tuple[bool, List[str]]:
        """
        Validate that required comparison columns exist
        
        Args:
            headers: Dictionary of column headers
            filename: Name of file being validated
        
        Returns:
            Tuple of (is_valid, missing_columns)
        """
        missing = []
        for col in config.COMPARISON_COLUMNS:
            if col not in headers:
                missing.append(col)
        
        return len(missing) == 0, missing
    
    def load_data(self) -> Tuple[bool, str]:
        """
        Load data from both Excel files
        
        Returns:
            Tuple of (success, error_message)
        """
        try:
            # Extract headers
            headers1 = self.extract_headers(self.ws1)
            headers2 = self.extract_headers(self.ws2)
            
            # Validate columns
            valid1, missing1 = self.validate_columns(headers1, self.file1_name)
            valid2, missing2 = self.validate_columns(headers2, self.file2_name)
            
            if not valid1:
                return False, f"File 1 missing columns: {', '.join(missing1)}"
            if not valid2:
                return False, f"File 2 missing columns: {', '.join(missing2)}"
            
            # Load data from file 1
            self.data1 = self._load_worksheet_data(self.ws1, headers1)
            
            # Load data from file 2
            self.data2 = self._load_worksheet_data(self.ws2, headers2)
            
            return True, ""
        
        except Exception as e:
            return False, f"Error loading data: {str(e)}"
    
    def _load_worksheet_data(self, worksheet: Any, headers: Dict[str, int]) -> Dict[Tuple, Dict]:
        """
        Load data from worksheet into dictionary
        
        Args:
            worksheet: openpyxl worksheet
            headers: Column header mapping
        
        Returns:
            Dictionary keyed by (CAD OEM Part Number, CAD OEM Rev)
        """
        data = {}
        
        for row_idx in range(2, worksheet.max_row + 1):
            # Get key columns
            part_num_col = headers.get('CAD OEM Part Number')
            rev_col = headers.get('CAD OEM Rev')
            
            if not part_num_col or not rev_col:
                continue
            
            part_num = worksheet.cell(row_idx, part_num_col).value
            rev = worksheet.cell(row_idx, rev_col).value
            
            # Skip rows without part number
            if not part_num:
                continue
            
            # Create key
            key = (str(part_num).strip() if part_num else '', str(rev).strip() if rev else '')
            
            # Extract all comparison columns
            row_data = {}
            for col_name in config.COMPARISON_COLUMNS:
                col_idx = headers.get(col_name)
                if col_idx:
                    cell_value = worksheet.cell(row_idx, col_idx).value
                    # Convert to string and strip, preserve None as empty string
                    row_data[col_name] = str(cell_value).strip() if cell_value is not None else ''
            
            # Store data (if duplicate keys exist, last one wins)
            data[key] = row_data
        
        return data
    
    def compare(self) -> Tuple[bool, str]:
        """
        Compare the two BOM files
        
        Returns:
            Tuple of (success, message)
        """
        try:
            # Find matching and unique keys
            keys1 = set(self.data1.keys())
            keys2 = set(self.data2.keys())
            
            self.matching_keys = sorted(list(keys1.intersection(keys2)))
            self.unique_file1 = sorted(list(keys1 - keys2))
            self.unique_file2 = sorted(list(keys2 - keys1))
            
            message = f"Comparison complete: {len(self.matching_keys)} matching, "
            message += f"{len(self.unique_file1)} unique in File 1, "
            message += f"{len(self.unique_file2)} unique in File 2"
            
            return True, message
        
        except Exception as e:
            return False, f"Error during comparison: {str(e)}"
    
    def _values_differ(self, val1: Any, val2: Any) -> bool:
        """
        Check if two values are different (including empty vs non-empty)
        
        Args:
            val1: First value
            val2: Second value
        
        Returns:
            True if values differ
        """
        # Convert to strings for comparison
        str1 = str(val1).strip() if val1 is not None else ''
        str2 = str(val2).strip() if val2 is not None else ''
        
        return str1 != str2
    
    def generate_comparison_report(self) -> BytesIO:
        """
        Generate Excel comparison report
        
        Returns:
            BytesIO buffer containing the Excel file
        """
        # Create new workbook
        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)
        
        # Sheet 1: Same (matching parts with differences highlighted)
        self._create_same_sheet(output_wb)
        
        # Sheet 2: Unique from File 1
        self._create_unique_sheet(output_wb, self.unique_file1, self.data1, 
                                  f"Unique_{self.sanitize_sheet_name(self.file1_name, 20)}")
        
        # Sheet 3: Unique from File 2
        self._create_unique_sheet(output_wb, self.unique_file2, self.data2, 
                                  f"Unique_{self.sanitize_sheet_name(self.file2_name, 20)}")
        
        # Save to BytesIO
        buffer = BytesIO()
        output_wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_same_sheet(self, workbook: Any):
        """Create the 'Same' sheet with matching parts and highlighted differences"""
        ws = workbook.create_sheet(title="Same")
        
        # Write headers
        for col_idx, col_name in enumerate(config.COMPARISON_COLUMNS, start=1):
            ws.cell(1, col_idx, col_name)
        
        # Write data from File 1 with highlighting
        row_idx = 2
        for key in self.matching_keys:
            data1 = self.data1[key]
            data2 = self.data2[key]
            
            for col_idx, col_name in enumerate(config.COMPARISON_COLUMNS, start=1):
                value1 = data1.get(col_name, '')
                value2 = data2.get(col_name, '')
                
                # Write value from File 1
                cell = ws.cell(row_idx, col_idx, value1)
                
                # Highlight if values differ (only for comparison columns, not key columns)
                if col_name not in ['CAD OEM Part Number', 'CAD OEM Rev']:
                    if self._values_differ(value1, value2):
                        cell.fill = self.yellow_fill
            
            row_idx += 1
    
    def _create_unique_sheet(self, workbook: Any, unique_keys: List[Tuple], 
                            data_dict: Dict[Tuple, Dict], sheet_name: str):
        """Create sheet for unique parts"""
        ws = workbook.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(config.COMPARISON_COLUMNS, start=1):
            ws.cell(1, col_idx, col_name)
        
        # Write data
        row_idx = 2
        for key in unique_keys:
            data = data_dict[key]
            
            for col_idx, col_name in enumerate(config.COMPARISON_COLUMNS, start=1):
                value = data.get(col_name, '')
                ws.cell(row_idx, col_idx, value)
            
            row_idx += 1
    
    def get_stats(self) -> Dict[str, int]:
        """
        Get comparison statistics
        
        Returns:
            Dictionary with comparison statistics
        """
        return {
            'total_file1': len(self.data1),
            'total_file2': len(self.data2),
            'matching': len(self.matching_keys),
            'unique_file1': len(self.unique_file1),
            'unique_file2': len(self.unique_file2)
        }
